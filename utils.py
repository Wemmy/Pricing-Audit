from collections.abc import Sequence
import openpyxl
import pandas as pd
import re
from openpyxl import load_workbook


def make_tuples(l):
    return tuple(make_tuples(i) if isinstance(i, Sequence) else i for i in l)

def read_table(file_name: str, table_name: str) -> pd.DataFrame:
    wb = openpyxl.load_workbook(file_name, read_only= False, data_only = True) # openpyxl does not have table info if read_only is True; data_only means any functions will pull the last saved value instead of the formula
    for sheetname in wb.sheetnames: # pulls as strings
        sheet = wb[sheetname] # get the sheet object instead of string
        if table_name in sheet.tables: # tables are stored within sheets, not within the workbook, although table names are unique in a workbook
            tbl = sheet.tables[table_name] # get table object instead of string
            tbl_range = tbl.ref #something like 'C4:F9'
            break # we've got our table, bail from for-loop
    data = sheet[tbl_range] # returns a tuple that contains rows, where each row is a tuple containing cells
    content = [[cell.value for cell in row] for row in data] # loop through those row/cell tuples
    header = content[0] # first row is column headers
    rest = content[1:] # every row that isn't the first is data
    df = pd.DataFrame(rest, columns = header)
    wb.close()
    return df

def get_max_index(singe_index_list):
    '''return all max index from a list of tuple'''
    max_sim = singe_index_list[0][1]
    list_index_sim = list(zip(*singe_index_list))
    list_index_max = []
    for i,s in enumerate(list_index_sim[1]):
        if s == max_sim:
            list_index_max.append(list_index_sim[0][i])
    return list_index_max

unit_list_1 = ['gm', 'g','gram','lb', 'gms', 'pounds', 'lbs', 'gr', 'kg', 'mg', 'litre','liter', 'loz','ml','oz', 'fl', 'ltr', 'l', 'lt', 'pkg','pk', 'ea', 'each', 'pcs', 'pc', 'btl', 'bottle','sl','cg', 'can' ,'sl']            
unit_list = ['gm', 'g','gram','lb', 'gms', 'pounds', 'lbs', 'gr', 'kg', 'mg', 'litre','liter', 'loz','ml','oz', 'fl', 'ltr', 'l', 'lt']

## data cleaning
def clean_size(doc):
    # lowercase the unit, 
    doc = doc.lower()
    # replace all special character with space
    doc = re.sub('[^a-zA-Z0-9\s]', ' ', doc)
    # remove the space before unit
    doc = re.sub(r'(\d+)(\s+)('+'|'.join(unit_list_1)+')(\s+|$)',r'\1\3\4', doc)
    # normalize the unit
    doc = re.sub(r'(\d+)(gr|gm|gram|gms)(\s+|$)',r'\1g\3', doc)
    doc = re.sub(r'(\d+)(lt|liter|litre|ltr)(\s+|$)',r'\1l\3', doc)
    doc = re.sub(r'(\d+)(pounds|lbs)(\s+|$)',r'\1lb\3', doc)
    doc = re.sub(r'(\d+)(pkg)(\s+|$)',r'\1pk\3', doc)
    doc = re.sub(r'(\d+)(each)(\s+|$)',r'\1ea\3', doc)
    doc = re.sub(r'(\d+)(pcs)(\s+|$)',r'\1pc\3', doc)
    doc = re.sub(r'(\d+)(bottle)(\s+|$)',r'\1btl\3', doc)
    return doc

def special_cases(doc):
    doc = re.sub('\s+root beer\s+',' rootbeer ', doc)
    doc = re.sub('\s+choc\s+',' chocolate ', doc)
    doc = re.sub('\s+snapd\s+',' snapped ', doc)
    doc = re.sub('\s+razzle berry\s+', ' razzleberry ', doc)
    doc = re.sub(r'\bsm\b',' small ', doc)
    doc = re.sub(r'\blg\b',' large ', doc)
    doc = re.sub(r'\bmd\b', ' medium ', doc)

    # some general rules:
    # 20oz -> LG; 16oz -> MD; 12oz ->SM; lg->large; md -> medium; sm->small
    # doc = re.sub(r'\bsmall\b',' 12oz ', doc)
    # doc = re.sub(r'\blarge\b',' 20oz ', doc)
    # doc = re.sub(r'\bmedium\b', ' medium ', doc)
    return doc

def split_size(doc):
    item = re.sub(r'\d+(?:'+ '|'.join(unit_list) +')(?:\s+|$)', '', doc).strip()
    size = re.findall(r'\d+(?:'+ '|'.join(unit_list) +')(?:\s+|$)', doc)
    # single space seperated
    item = re.sub(r"\s+", " ", item)
    
    if len(size) > 1:
        size = [size[0]]
    if not size:
        size = ['']
    return [item] + size

def data_cleaning(doc):
    return split_size(special_cases(clean_size(doc)))

def read_excel_file(file_name, sheet_name, header_row = 0):

    # Load the Excel file
    wb = load_workbook(filename=file_name, read_only=True, data_only=True)

    # Get the sheet
    sheet = wb[sheet_name]

    # Create a pandas DataFrame from the sheet data
    df = pd.DataFrame(sheet.values)

    # Set the headers to the first row of the DataFrame
    header_cells = sheet[header_row]
    headers = [cell.value for cell in header_cells]

    # Create a pandas DataFrame from the sheet data
    rows = sheet.iter_rows(min_row=header_row+1)
    data = [[cell.value for cell in row] for row in rows]
    df = pd.DataFrame(data, columns=headers)

    # Close the workbook
    wb.close()

    return df



if __name__ == '__main__':
    pass